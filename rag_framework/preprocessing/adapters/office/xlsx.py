"""Adapter openpyxl pour les fichiers Excel (.xlsx).

Auteur: RAG Framework Team
Version: 1.0.0
"""

from typing import Any, ClassVar

from rag_framework.preprocessing.adapters.base import LibraryAdapter, ParsingError


class OpenpyxlAdapter(LibraryAdapter):
    """Adapter pour la librairie openpyxl.

    Extrait le texte des fichiers Excel (.xlsx).

    Attributes:
        REQUIRED_MODULES: Liste des modules requis.
    """

    REQUIRED_MODULES: ClassVar[list[str]] = ["openpyxl"]

    def parse(self, file_path: str) -> dict[str, Any]:
        """Parse un fichier Excel avec openpyxl.

        Args:
            file_path: Chemin vers le fichier .xlsx.

        Returns:
            Dictionnaire avec text, metadata, sheets.

        Raises:
            ParsingError: Si le parsing échoue.
        """
        try:
            from openpyxl import load_workbook

            # Configuration
            read_only = self.library_config.get("read_only", True)
            data_only = self.library_config.get("data_only", True)

            # Ouvrir le workbook
            wb = load_workbook(file_path, read_only=read_only, data_only=data_only)

            # Extraire le texte de toutes les feuilles
            sheets = []
            full_text = []

            for sheet in wb.worksheets:
                sheet_dict: dict[str, Any] = {
                    "name": sheet.title,
                    "rows": [],
                    "text": "",
                }

                # Extraire toutes les lignes
                sheet_text = []
                for row in sheet.iter_rows(values_only=True):
                    # Filtrer les valeurs None et convertir en strings
                    row_values = [str(cell) for cell in row if cell is not None]
                    if row_values:
                        row_text = " | ".join(row_values)
                        sheet_dict["rows"].append(row_values)
                        sheet_text.append(row_text)

                sheet_dict["text"] = "\n".join(sheet_text)
                sheet_dict["row_count"] = len(sheet_dict["rows"])
                sheets.append(sheet_dict)
                full_text.extend(sheet_text)

            # Métadonnées
            metadata = {
                "sheet_count": len(sheets),
                "sheet_names": [s["name"] for s in sheets],
                "total_rows": sum(s["row_count"] for s in sheets),
            }

            # Propriétés du document si disponibles
            if hasattr(wb, "properties"):
                props = wb.properties
                metadata.update(
                    {
                        "title": props.title or "",
                        "creator": props.creator or "",
                        "created": str(props.created) if props.created else "",
                        "modified": str(props.modified) if props.modified else "",
                    }
                )

            wb.close()

            return {
                "text": "\n\n".join(full_text),
                "metadata": metadata,
                "sheets": sheets,
            }

        except Exception as e:
            raise ParsingError(f"Échec openpyxl parsing : {e}") from e
